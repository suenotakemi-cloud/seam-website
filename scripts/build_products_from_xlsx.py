#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SEAM 商品データ JSON 生成スクリプト
=====================================
入力 : /Users/suenotakemi/Downloads/FWBUZR7u.xlsx
出力 : data/products-from-xlsx.json

固定タグ語彙（CLAUDE.md準拠）に厳密にマッピングし、
公式 feature_points / recommended_for / ライン情報から
診断エンジン用のタグと物語フィールドを生成する。
香り情報は出力しない。価格・サイズは xlsx に無いため省略。
"""

import openpyxl
import json
import re
import unicodedata
from pathlib import Path

XLSX_PATH = "/Users/suenotakemi/Downloads/FWBUZR7u.xlsx"
OUT_PATH  = "/Users/suenotakemi/Downloads/code_sandbox_light_a3728b14_1778910042/data/products-from-xlsx.json"


# ============================================================
# 1) ブランド辞書（日本語表記 → 英語表示名 + メーカー）
# ============================================================
BRAND_MAP = {
    "HITA":              {"display": "HITA",              "slug": "hita",          "maker": "ルベル / タカラベルモント"},
    "LOA THE OIL":       {"display": "LOA THE OIL",       "slug": "loa",           "maker": "LOA"},
    "SEE/SAW":           {"display": "SEE/SAW",           "slug": "seesaw",        "maker": "アリミノ"},
    "TOKIO INKARAMI":    {"display": "TOKIO INKARAMI",    "slug": "tokio",         "maker": "TOKIO"},
    "アルピジョン":       {"display": "ALPSION",           "slug": "alpsion",       "maker": "アルピジョン"},
    "エルジューダ":       {"display": "Elujuda",           "slug": "elujuda",       "maker": "ミルボン"},
    "オッジィオット":     {"display": "Oggi otto",         "slug": "oggiotto",      "maker": "テクノエイト"},
    "オージュア":         {"display": "Aujua",             "slug": "aujua",         "maker": "ミルボン"},
    "グローバルミルボン": {"display": "Global Milbon",     "slug": "globalmilbon",  "maker": "ミルボン"},
    "ケラスターゼ":       {"display": "Kérastase",         "slug": "kerastase",     "maker": "ロレアル"},
    "サブリミック":       {"display": "SUBLIMIC",          "slug": "sublimic",      "maker": "資生堂プロフェッショナル"},
    "スプリナージュ":     {"display": "SPRINAGE",          "slug": "sprinage",      "maker": "アリミノ"},
    "トラック":           {"display": "track",             "slug": "track",         "maker": "ナンバースリー"},
    "ナノアミノ":         {"display": "NANOAMINO",         "slug": "nanoamino",     "maker": "ニューウェイジャパン"},
    "ハホニコ":           {"display": "HAHONICO",          "slug": "hahonico",      "maker": "ハホニコ"},
    "バイカルテ":         {"display": "Vie qualite",       "slug": "viequalite",    "maker": "アリミノ"},
    "リケラ":             {"display": "LIKERA",            "slug": "likera",        "maker": "リトルサイエンティスト"},
    "髪にドラマを。":     {"display": "髪にドラマを。",     "slug": "drama",         "maker": "髪にドラマを。"},
}


# ============================================================
# 2) ライン辞書（日本語ライン名 → 英語 slug）
#    (brand_jp, line_jp) で一意。同じライン名でも分けたいときに使う。
# ============================================================
LINE_SLUG = {
    # HITA
    ("HITA", "トリートメント"):                 "hita",
    ("HITA", "ヒタ オイル"):                     "oil",
    # LOA
    ("LOA THE OIL", "LOA THE OIL"):              "core",
    # SEE/SAW
    ("SEE/SAW", "B"):                            "balance",
    ("SEE/SAW", "S"):                            "smooth",
    ("SEE/SAW", "T"):                            "tight",
    ("SEE/SAW", "ヘアメイク"):                   "hairmake",
    # TOKIO
    ("TOKIO INKARAMI", "TOKIO IE INKARAMI HEAD SPA"):  "headspa",
    ("TOKIO INKARAMI", "リミテッド"):                   "limited",
    ("TOKIO INKARAMI", "TOKIO IE INKARAMI PLATINUM"):   "platinum",
    ("TOKIO INKARAMI", "TOKIO IE INKARAMI PREMIUM"):    "premium",
    ("TOKIO INKARAMI", "アウトカラミ"):                  "outkarami",
    # ALPSION
    ("アルピジョン", "ALPSION HAIR CARE"):       "core",
    # Elujuda
    ("エルジューダ", "FO"):                      "fo",
    ("エルジューダ", "エマルジョン+"):           "emulsion",
    ("エルジューダ", "エマルジョン"):            "graceon",
    ("エルジューダ", "フリッズフィクサー エマルジョン"): "frizzfixer",
    ("エルジューダ", "ジェルセラム"):            "bleachcare",
    # Oggi otto
    ("オッジィオット", "スキャルプ"):            "scalp",
    ("オッジィオット", "セラム"):                "serum",
    ("オッジィオット", "アウトバス"):            "outbath",
    # Aujua
    ("オージュア", "アクアヴィア"):              "aquavia",
    ("オージュア", "イミュライズ"):              "immurise",
    ("オージュア", "インメトリィ"):              "inmetry",
    ("オージュア", "エイジングスパ"):            "agingspa",
    ("オージュア", "クエンチ"):                  "quench",
    ("オージュア", "スムース"):                  "smooth",
    ("オージュア", "フィルメロウ"):              "filmellow",
    ("オージュア", "モイストカーム"):            "moistcalm",
    ("オージュア", "リペアリティ"):              "repairity",
    # Global Milbon
    ("グローバルミルボン", "アンチフリッズ"):    "antifrizz",
    ("グローバルミルボン", "ブロンドプラス"):    "blondeplus",
    ("グローバルミルボン", "カラープリザーブ"):  "colorpreserve",
    ("グローバルミルボン", "モイスチュア"):      "moisture",
    ("グローバルミルボン", "リペア"):            "repair",
    ("グローバルミルボン", "リペアヒート"):      "repairheat",
    ("グローバルミルボン", "スムース"):          "smooth",
    # Kérastase
    ("ケラスターゼ", "ブロンドアブソリュ"):      "blondabsolu",
    ("ケラスターゼ", "クロマアブソリュ"):        "chromaabsolu",
    ("ケラスターゼ", "クロノロジスト"):          "chronologiste",
    ("ケラスターゼ", "ディシプリン"):            "discipline",
    ("ケラスターゼ", "ジェネシス"):              "genesis",
    ("ケラスターゼ", "ニュートリティブ"):        "nutritive",
    ("ケラスターゼ", "レジスタンス"):            "resistance",
    # SUBLIMIC
    ("サブリミック", "アクアインテンシブ"):      "aquaintensive",
    ("サブリミック", "アデノバイタル"):          "adenovital",
    ("サブリミック", "エアリーフロー"):          "airyflow",
    ("サブリミック", "フェンテフォルテ"):        "fuenteforte",
    ("サブリミック", "ルミノフォース"):          "luminoforce",
    ("サブリミック", "ワンダーシールド"):        "wondershield",
    # SPRINAGE
    ("スプリナージュ", "SPRINAGE HAIR CARE"):    "core",
    # track
    ("トラック", "track Cream"):                 "cream",
    # NANOAMINO
    ("ナノアミノ", "NANOAMINO HOMECARE"):        "homecare",
    # HAHONICO
    ("ハホニコ", "キラメラメ系"):                "ramerame",
    # Vie qualite
    ("バイカルテ", "リペアシャンプー CH+"):      "ch",
    ("バイカルテ", "リペアシャンプー FH+"):      "fh",
    ("バイカルテ", "リペアシャンプー UH+"):      "uh",
    ("バイカルテ", "セラムトリートメント SS+/MS+/HS+"): "serum",
    ("バイカルテ", "コンセントレイトエッセンス"): "concentrate",
    ("バイカルテ", "セラムマスク"):              "serummask",
    ("バイカルテ", "エッセンスミルク"):          "essence",
    # LIKERA
    ("リケラ", "リケラエマルジョン"):            "emulsion",
    # 髪にドラマを。
    ("髪にドラマを。", "つるりんちょ。シャンプー"):                 "tsurun-sh",
    ("髪にドラマを。", "つるりんちょ。ブースター"):                 "tsurun-booster",
    ("髪にドラマを。", "つるりんちょ。トリートメント"):             "tsurun-tr",
    ("髪にドラマを。", "つるりんちょ。トリートメント SARARITO"):    "tsurun-sararito",
    ("髪にドラマを。", "いるかのせなか。ベースケア"):                "iruka",
}


# ============================================================
# 3) カテゴリ判定（xlsx category + product name → SEAM category）
# ============================================================
def detect_category(cat_jp: str, name: str):
    """
    SEAM標準カテゴリと routineStep を返す。
    アウトバスは name から oil / milk / styling を分岐。
    """
    name_l = name.lower()
    if cat_jp == "シャンプー":
        return "shampoo", 1, "sh"
    if cat_jp == "トリートメント":
        return "treatment", 2, "tr"
    if cat_jp == "マスク":
        return "hair-mask", 3, "mask"
    if cat_jp == "エッセンス":
        # ほぼ頭皮系
        return "scalp-essence", 6, "essence"
    if cat_jp == "アウトバス":
        # 名前から細分化
        if any(k in name for k in ["バーム", "Balm", "balm"]):
            return "styling", 7, "balm"
        if any(k in name for k in ["オイル", "Oil", "OIL"]):
            return "out-bath-oil", 5, "oil"
        if any(k in name for k in ["ミスト", "Mist"]):
            return "out-bath-milk", 4, "mist"
        if any(k in name for k in ["ジェル", "Gel"]):
            return "out-bath-milk", 4, "gel"
        if any(k in name for k in ["スカルプエッセンス", "頭皮用美容液", "スカルプ・毛髪ケア", "スカルプ"]):
            return "scalp-essence", 6, "essence"
        if any(k in name for k in ["ローション", "Lotion"]):
            return "scalp-essence", 6, "lotion"
        # ミルク / エマルジョン / セラム / クリーム
        return "out-bath-milk", 4, "out"
    # フォールバック
    return "treatment", 2, "tr"


# ============================================================
# 4) feature_points / recommended_for / product context から
#    concernTags / hairType / damageTags / finishTags /
#    functionTags / scalpFit / preferenceFit / usageTiming /
#    specialTech を導く
# ============================================================
def derive_tags(brand_jp, line_jp, name, cat_jp, points, recommend, category, role):
    """
    返り値: dict ベースのタグ集合
    """
    text = " ".join([str(points or ""), str(recommend or ""), str(role or ""), str(name or ""), str(line_jp or "")]).lower()
    points_l = (points or "").lower()
    rec_l    = (recommend or "").lower()

    tags = {
        "concernTags": [],
        "hairType":    {"thickness": [], "density": [], "wave": []},
        "damageTags":  {},
        "finishTags":  {"finish": [], "texture": []},
        "functionTags": [],
        "scalpFit":    [],
        "lengthFit":   "all",
        "lifestyleFit": [],
        "preferenceFit": [],
        "usageTiming": [],
        "specialTech": [],
    }

    # ---------- concernTags ----------
    def add_concern(*keys):
        for k in keys:
            if k not in tags["concernTags"]:
                tags["concernTags"].append(k)

    if any(k in text for k in ["乾燥", "保湿", "水分保持", "うるおい"]):
        add_concern("dry")
    if any(k in text for k in ["パサつき", "パサ"]):
        add_concern("frizz", "dry")
    if any(k in text for k in ["うねり", "くせ", "湿気"]):
        add_concern("wave")
    if any(k in text for k in ["ごわつき", "ゴワ"]):
        add_concern("rough")
    if any(k in text for k in ["絡まり", "もつれ"]):
        add_concern("tangle")
    if any(k in text for k in ["ツヤがない", "ツヤ不足"]):
        add_concern("no-shine")
    if "枝毛" in text or "切れ毛" in text:
        add_concern("split")
    if any(k in text for k in ["ダメージ", "ハイダメージ", "補修", "ブリーチ", "高補修"]):
        add_concern("damage")
    if any(k in text for k in ["カラーケア", "色持ち", "褪色", "色落ち", "カラー後"]):
        add_concern("color-fade")
    if any(k in text for k in ["白髪", "グレイ"]):
        add_concern("gray-fade")
    if any(k in text for k in ["ボリュームダウン", "重くしたくない", "軽い", "軽やか"]):
        # 「軽い仕上がり」は volume-down ではなく preferenceFit へ
        pass
    if any(k in text for k in ["広がり", "膨らみ"]):
        add_concern("spread")
    if "まとまらない" in text or "まとまり" in text:
        # まとまりが課題のときだけ no-smooth を当てる（くせ・うねり系はほぼ該当）
        if any(k in text for k in ["くせ", "うねり", "広がり", "まとまらない"]):
            add_concern("no-smooth")
    if any(k in text for k in ["頭皮の乾燥", "頭皮乾燥"]):
        add_concern("scalp-dry")
    if any(k in text for k in ["べたつき", "ベタつき"]):
        add_concern("scalp-oily")
    if any(k in text for k in ["におい", "ニオイ"]):
        add_concern("scalp-odor")
    if "かゆみ" in text:
        add_concern("scalp-itch")
    if any(k in text for k in ["抜け毛", "薄毛", "育毛"]):
        add_concern("thinning")
    if any(k in text for k in ["熱ダメージ", "アイロン", "コテ", "ヒート"]):
        add_concern("heat-damage")
    if any(k in text for k in ["光沢", "ツヤ"]):
        # 光沢=ツヤ — ただし「ツヤがない」と書かれていなければ shine-loss は付けない
        pass
    if any(k in text for k in ["ハリ", "コシ", "ハリコシ", "弱り"]):
        add_concern("no-firmness")
    if any(k in text for k in ["トップのふんわり", "根元ボリューム", "ふんわり"]) and "ボリューム" in text:
        add_concern("top-flat", "volume-down")

    # ---------- hairType ----------
    thickness = set()
    density   = set()
    wave      = set()
    if any(k in text for k in ["細毛", "軟毛", "細い"]):
        thickness.add("thin")
        density.add("low")
    if any(k in text for k in ["硬毛", "硬い"]):
        thickness.add("thick")
        density.add("high")
    if "多毛" in text or "膨らみ" in text:
        density.add("high")
    if any(k in text for k in ["普通毛", "ノーマル", "全髪質"]):
        thickness.add("normal")
        density.add("normal")
    if any(k in text for k in ["くせ", "うねり", "縮毛"]):
        wave.update({"medium", "strong"})
    if "縮毛矯正毛" in text or "矯正毛" in text:
        wave.add("straightened")
    if any(k in text for k in ["直毛", "ストレート"]):
        wave.add("none")
    # 何も決まらなければ全方位対応とみなす
    if not thickness:
        thickness = {"thin", "normal", "thick"}
    if not density:
        density = {"low", "normal", "high"}
    if not wave:
        wave = {"none", "weak", "medium"}
    tags["hairType"]["thickness"] = sorted(thickness, key=["thin","normal","thick"].index)
    tags["hairType"]["density"]   = sorted(density,   key=["low","normal","high"].index)
    tags["hairType"]["wave"]      = sorted(wave,      key=["none","weak","medium","strong","straightened"].index)

    # ---------- damageTags ----------
    # bleachOk
    if "ブリーチ" in text or "ハイトーン" in text or "ハイダメージ" in text:
        tags["damageTags"]["bleachOk"] = "heavy"
    elif "ダメージ補修" in text or "高補修" in text or "ダメージ" in text:
        tags["damageTags"]["bleachOk"] = "medium"
    else:
        tags["damageTags"]["bleachOk"] = "light"
    # colorCare
    if any(k in text for k in ["カラーケア", "色持ち", "褪色", "color"]):
        tags["damageTags"]["colorCare"] = "fade-prevent"
    else:
        tags["damageTags"]["colorCare"] = "maintain"
    # straightenOk
    if "縮毛矯正" in text or "ストレート毛" in text or "矯正毛" in text:
        tags["damageTags"]["straightenOk"] = "mid"
    else:
        tags["damageTags"]["straightenOk"] = "new"
    # permOk
    if "パーマ" in text or "デジタル" in text:
        tags["damageTags"]["permOk"] = "digital"
    else:
        tags["damageTags"]["permOk"] = "normal"
    # heatTolerance / heatProtect
    if any(k in text for k in ["ヒートプロテクト", "熱ダメージ", "アイロン", "コテ", "リペアヒート"]):
        tags["damageTags"]["heatTolerance"] = "t200"
        tags["damageTags"]["heatProtect"]   = True
    else:
        tags["damageTags"]["heatTolerance"] = "t180"
        tags["damageTags"]["heatProtect"]   = False

    # ---------- finishTags ----------
    finish = set()
    weight = "medium"
    texture = set()
    if any(k in text for k in ["しっとり", "保湿", "うるおい", "moist"]):
        finish.add("moist")
    if any(k in text for k in ["ツヤ", "艶", "glossy", "光沢"]):
        finish.add("glossy")
    if any(k in text for k in ["ハリ", "コシ", "ふんわり", "ボリューム"]):
        finish.add("firm")
    if any(k in text for k in ["なめらか", "smooth", "指通り"]):
        finish.add("smooth")
    if any(k in text for k in ["軽やか", "さらさら", "airy", "ふわっと"]):
        finish.add("airy")
        weight = "light"
    if any(k in text for k in ["やわらかさ", "soft", "やわらか"]):
        finish.add("soft")
    if "しっとり" in text and "軽" not in text:
        weight = "heavy" if "濃密" in text or "重め" in text else "medium"
    # texture
    if any(k in text for k in ["silky", "シルキー", "なめらか"]):
        texture.add("silky")
    if any(k in text for k in ["ふんわり", "ボリューム"]):
        texture.add("volume")
    if any(k in text for k in ["bonding", "ボンド", "結合", "インカラミ"]):
        texture.add("bonding")
    if any(k in text for k in ["自然", "ナチュラル", "natural"]):
        texture.add("natural")
    if not finish:
        finish.add("smooth")
    if not texture:
        texture.add("silky")
    tags["finishTags"]["finish"]  = sorted(finish,  key=["moist","glossy","firm","smooth","airy","soft"].index)
    tags["finishTags"]["weight"]  = weight
    tags["finishTags"]["texture"] = sorted(texture, key=["silky","volume","bonding","natural"].index)

    # ---------- functionTags ----------
    fn = set()
    if any(k in text for k in ["保湿", "うるおい", "水分", "moist"]):
        fn.add("moisture")
    if any(k in text for k in ["内部補修", "補修", "ダメージ補修", "高補修"]):
        fn.add("internal-repair")
    if any(k in text for k in ["表面補修", "表面"]):
        fn.add("surface-repair")
    if any(k in text for k in ["ボンド", "結合", "インカラミ", "プレックス", "plex"]):
        fn.add("bonding")
    if any(k in text for k in ["ヒートプロテクト", "熱ダメージ", "アイロン", "コテ"]):
        fn.add("heat-protect")
    if "uv" in text or "紫外線" in text:
        fn.add("uv-protect")
    if any(k in text for k in ["カラーケア", "色持ち", "褪色", "カラーキープ"]):
        fn.add("color-lock")
    if any(k in text for k in ["酸熱", "ph調整"]):
        fn.add("ph-balance")
    if any(k in text for k in ["ボリューム", "ふんわり", "ふんわりトップ"]):
        fn.add("volume-up")
    if any(k in text for k in ["ディープクレンジング", "クレンジング"]):
        fn.add("clarify")
    if any(k in text for k in ["頭皮保湿"]):
        fn.add("scalp-moisture")
    if any(k in text for k in ["頭皮ケア", "地肌", "スキャルプ", "スカルプ"]):
        fn.add("scalp-clean")
    if any(k in text for k in ["エイジング", "ハリコシ", "ボリュームダウン", "薄毛", "抜け毛", "アデノバイタル", "イミュライズ", "agingspa"]):
        fn.add("aging-care")
    tags["functionTags"] = sorted(fn)

    # ---------- scalpFit ----------
    if any(k in text for k in ["頭皮", "スキャルプ", "スカルプ", "地肌"]):
        scalp = set()
        if any(k in text for k in ["乾燥", "敏感"]):
            scalp.add("dry-scalp")
        if any(k in text for k in ["ベタつき", "べたつき", "脂性"]):
            scalp.add("oily-scalp")
        if any(k in text for k in ["敏感", "ゆらぎ"]):
            scalp.add("sensitive")
        if any(k in text for k in ["エイジング", "ハリコシ", "薄毛", "抜け毛"]):
            scalp.add("aging-scalp")
        if not scalp:
            scalp.add("normal")
        tags["scalpFit"] = sorted(scalp)

    # ---------- preferenceFit ----------
    pref = set()
    if any(k in text for k in ["軽い", "軽やか", "さらさら", "重くしたくない", "軽い仕上がり"]):
        pref.add("weight-light")
    if any(k in text for k in ["ベタつき", "べたつき", "オイル感を抑え"]):
        pref.add("non-oily")
    if any(k in text for k in ["持続", "長持ち", "サロン仕上がり"]):
        pref.add("long-lasting")
    tags["preferenceFit"] = sorted(pref)

    # ---------- usageTiming ----------
    ut = set()
    if category in ("shampoo", "treatment"):
        ut.add("daily-in-bath")
    if category == "hair-mask":
        ut.add("weekly-mask")
    if category in ("out-bath-milk", "out-bath-oil", "styling"):
        ut.add("before-dry")
        if tags["damageTags"].get("heatProtect"):
            ut.add("heat-before")
    if category == "scalp-essence":
        ut.add("night-scalp")
    tags["usageTiming"] = sorted(ut)

    # ---------- specialTech ----------
    st = set()
    if brand_jp == "TOKIO INKARAMI":
        st.add("inkarami")
        if any(k in text for k in ["プラチナム", "リミテッド", "ハイダメージ"]):
            st.add("plex-bonding")
    if brand_jp == "オージュア":
        st.add("aujua-system")
    if "酸熱" in text:
        st.add("acid-heat")
    if any(k in text for k in ["セラミド", "ceramide"]):
        st.add("ceramide")
    if any(k in text for k in ["ケラチン", "keratin"]):
        st.add("keratin")
    if "cmc" in text:
        st.add("cmc")
    if any(k in text for k in ["ヒアルロン", "hyaluronic"]):
        st.add("hyaluronic")
    if any(k in text for k in ["植物", "ボタニカル", "自然派", "botanical"]):
        st.add("botanical")
    if any(k in text for k in ["プレックス", "plex", "ボンド"]):
        st.add("plex-bonding")
    tags["specialTech"] = sorted(st)

    # 空フィールド整理（あとで dict に格納する側で省略判定）
    return tags


# ============================================================
# 5) priceTier は xlsx に無いので省略する（推測しない原則）
# ============================================================


# ============================================================
# 6) 物語フィールド（pitchCopy / targetHair / targetPerson /
#    solvesConcern / effects / usage / keyIngredients）
#    ラインごとにテンプレ → カテゴリで微調整
# ============================================================
LINE_NARRATIVE = {
    # ---------- HITA ----------
    ("HITA", "トリートメント"): {
        "pitchCopy":     "くせ・うねりの広がりに、まとまりの土台を。",
        "targetHair":    "くせ・うねりで広がりやすく、日々まとまりにくい髪。",
        "targetPerson":  "「朝、髪が広がってまとまらない」「湿気で扱いにくくなる」と感じている方。",
        "solvesConcern": "うねり・くせ毛にアプローチする処方で、毎日のケアから広がりを抑え、扱いやすい質感へ整えます。",
        "keyIngredients": [],
        "effects": [
            "くせ・うねりによる広がりを整える",
            "なめらかな指通りとまとまりを与える",
            "日常で扱いやすい質感へ導く",
        ],
    },
    ("HITA", "ヒタ オイル"): {
        "pitchCopy":     "広がりを整え、上品なツヤを足すアウトバスオイル。",
        "targetHair":    "広がりやパサつきが気になる髪。ツヤを足したい仕上げ前の髪。",
        "targetPerson":  "「広がりを抑えながら自然なツヤがほしい」方。",
        "solvesConcern": "毛先まで均一になじむオイルが広がりを抑え、しっとりとしたツヤとまとまりを与えます。",
        "keyIngredients": [],
        "effects": [
            "毛先のパサつきと広がりを抑える",
            "ツヤとなめらかな指通りを与える",
            "仕上げの質感を整える",
        ],
        "usage": "タオルドライ後、毛先〜中間に。仕上げにも少量を。",
    },
    # ---------- LOA ----------
    ("LOA THE OIL", "LOA THE OIL"): {
        "pitchCopy":     "仕上げを格上げするマルチユースオイル。",
        "targetHair":    "ツヤとまとまりを足したい全髪質。",
        "targetPerson":  "「サロン帰りのような上質な仕上げを毎日続けたい」方。",
        "solvesConcern": "髪・肌・全身に使えるマルチオイル。毛先になじませることで上質なツヤと指通りを与えます。",
        "keyIngredients": [],
        "effects": [
            "毛先のパサつきを抑える",
            "上質なツヤと指通りを与える",
            "ヘア・スキンに使えるマルチユース",
        ],
        "usage": "タオルドライ後、または乾いた髪の仕上げに毛先中心に。",
    },
    # ---------- SEE/SAW ----------
    ("SEE/SAW", "B"): {
        "pitchCopy":     "ツヤとやわらかさのバランスを取る、定番ライン。",
        "targetHair":    "髪と頭皮の両方を整えたい、ノーマル毛全般。",
        "targetPerson":  "「ツヤとやわらかさを両立したい」方。",
        "solvesConcern": "ヘア&スキャルプ設計が、髪と頭皮の両方をなめらかに整え、自然なツヤとやわらかさを引き出します。",
        "keyIngredients": [],
        "effects": [
            "髪と頭皮を同時にケアする",
            "自然なツヤとやわらかさを与える",
            "ベーシックなまとまりを整える",
        ],
    },
    ("SEE/SAW", "S"): {
        "pitchCopy":     "細毛・絡まりに、軽やかなさらさらの仕上がりを。",
        "targetHair":    "細毛・軟毛・絡まりが気になる、軽い仕上がりを好む髪。",
        "targetPerson":  "「重くなりたくない」「さらっとした指通りが好き」な方。",
        "solvesConcern": "軽やかな仕上がりを叶える設計で、細毛や絡まりがちな髪をふんわりと整え、透明感のあるさらさら感を与えます。",
        "keyIngredients": [],
        "effects": [
            "絡まりを抑え、軽い指通りに整える",
            "細毛・軟毛をふんわり仕上げる",
            "透明感のあるさらさら感を与える",
        ],
    },
    ("SEE/SAW", "T"): {
        "pitchCopy":     "広がり・ボリュームに、しっかりまとまりを与える一本。",
        "targetHair":    "広がり・ボリュームが気になり、まとめにくい髪。",
        "targetPerson":  "「ボリュームを抑えてすっきりまとめたい」方。",
        "solvesConcern": "広がりを抑える設計が髪を内側からしっとり整え、まとまりとツヤを両立します。",
        "keyIngredients": [],
        "effects": [
            "広がり・ボリュームを抑える",
            "まとまりとなめらかな指通りを与える",
            "落ち着いたツヤを引き出す",
        ],
    },
    ("SEE/SAW", "ヘアメイク"): {
        "pitchCopy":     "仕上げの質感を、自由に操る一本。",
        "targetHair":    "仕上げのツヤや束感を整えたい全髪質。",
        "targetPerson":  "「仕上げの質感までこだわりたい」方。",
        "solvesConcern": "ヘアメイク発想の処方が、束感・ツヤ・透明感を自在にコントロールします。",
        "keyIngredients": [],
        "effects": [
            "仕上げのツヤと透明感を整える",
            "束感・動きを自然に演出する",
            "スタイリングの幅を広げる",
        ],
        "usage": "乾いた髪の仕上げに、毛先中心に少量。",
    },
    # ---------- TOKIO INKARAMI ----------
    ("TOKIO INKARAMI", "TOKIO IE INKARAMI HEAD SPA"): {
        "pitchCopy":     "頭皮ケアとヘアケアを、一本でつなげる。",
        "targetHair":    "頭皮環境とダメージを同時に整えたい髪。",
        "targetPerson":  "「ヘアケアと頭皮ケアを両立したい」方。",
        "solvesConcern": "頭皮ケアに着目したインカラミ処方で、地肌を整えながら毛髪内部まで補修します。",
        "keyIngredients": ["インカラミ処方"],
        "effects": [
            "頭皮環境を整え、すっきりとした洗い心地",
            "髪内部までしっかり補修",
            "ヘアケアと頭皮ケアを同時に",
        ],
    },
    ("TOKIO INKARAMI", "リミテッド"): {
        "pitchCopy":     "TOKIOの最上級補修。ハリ感としっとりまとまりを同時に。",
        "targetHair":    "繰り返したダメージで弱り、ハリも欲しい髪。",
        "targetPerson":  "「補修感も上質なまとまりも妥協したくない」方。",
        "solvesConcern": "プレミアム配合のインカラミ処方が、ハイダメージ毛にも補修感としっとりまとまりを与えます。",
        "keyIngredients": ["インカラミ処方"],
        "effects": [
            "ハイダメージ毛の内部結合を補修",
            "ハリ感としっとりまとまりを両立",
            "上質なツヤを与える",
        ],
    },
    ("TOKIO INKARAMI", "TOKIO IE INKARAMI PLATINUM"): {
        "pitchCopy":     "重くしないのに、毛先までしなやかに。",
        "targetHair":    "ダメージを補修しつつ、軽さも欲しい髪。",
        "targetPerson":  "「補修はしたい、でも重くなりたくない」方。",
        "solvesConcern": "プラチナム配合のインカラミ処方が、ダメージを補修しながらさらさらの軽いツヤを引き出します。",
        "keyIngredients": ["プラチナム配合インカラミ処方"],
        "effects": [
            "ダメージ毛を補修しながら軽さを残す",
            "さらさらの指通りと上品なツヤを与える",
            "重くなりがちな髪も扱いやすく整える",
        ],
    },
    ("TOKIO INKARAMI", "TOKIO IE INKARAMI PREMIUM"): {
        "pitchCopy":     "乾燥と広がりに、しっとりまとまりの補修を。",
        "targetHair":    "乾燥・広がり・パサつきが気になるダメージ毛。",
        "targetPerson":  "「乾燥や広がりを抑え、しっとりまとめたい」方。",
        "solvesConcern": "プレミアム処方のインカラミが、乾燥した髪の内部を補修し、しっとりとしたまとまりを引き出します。",
        "keyIngredients": ["インカラミ処方"],
        "effects": [
            "乾燥・パサつきによる広がりを抑える",
            "しっとりとしたまとまりを与える",
            "ダメージ補修と保湿を両立",
        ],
    },
    ("TOKIO INKARAMI", "アウトカラミ"): {
        "pitchCopy":     "ドライヤー前の補修ケアという、新しい一手。",
        "targetHair":    "毎日の熱・乾燥ダメージが気になる髪。",
        "targetPerson":  "「ドライヤー前にもしっかり補修したい」方。",
        "solvesConcern": "アウトバスでも内部補修にこだわるインカラミ発想。乾かす前のひと手間で、ツヤと指通りを引き上げます。",
        "keyIngredients": ["インカラミ処方"],
        "effects": [
            "ドライ前の髪を補修し、ツヤと指通りを高める",
            "毛先のパサつき・絡まりを整える",
            "熱・乾燥ダメージから髪を守る",
        ],
        "usage": "タオルドライ後、毛先〜中間に。ドライヤーの前に。",
    },
    # ---------- ALPSION ----------
    ("アルピジョン", "ALPSION HAIR CARE"): {
        "pitchCopy":     "やさしい使用感で、毎日の髪をいたわるヘアケア。",
        "targetHair":    "刺激感を抑えた日常ケアを求める髪。",
        "targetPerson":  "「やさしい使い心地で、毎日続けやすいケアがいい」方。",
        "solvesConcern": "やさしい使用感の処方が、髪と頭皮をいたわりながら、日々のうるおいとまとまりを整えます。",
        "keyIngredients": [],
        "effects": [
            "髪と頭皮をやさしくいたわる",
            "うるおいとまとまりを与える",
            "毎日続けやすい使用感",
        ],
    },
    # ---------- Elujuda ----------
    ("エルジューダ", "FO"): {
        "pitchCopy":     "細毛・普通毛に、軽やかなツヤとまとまりを。",
        "targetHair":    "細毛〜普通毛で、軽さとツヤを両立したい髪。",
        "targetPerson":  "「重くなりたくないけれど、ツヤがほしい」方。",
        "solvesConcern": "軽やかなオイル設計が、髪に負担をかけずにツヤとまとまりを与えます。",
        "keyIngredients": [],
        "effects": [
            "軽い質感のままツヤを与える",
            "細毛・普通毛のまとまりを整える",
            "毛先まで指通りをなめらかに",
        ],
        "usage": "タオルドライ後、毛先〜中間になじませる。",
    },
    ("エルジューダ", "エマルジョン+"): {
        "pitchCopy":     "硬毛・広がりに、しっとり寄せるエマルジョン。",
        "targetHair":    "硬毛・多毛・広がり・パサつきが気になる髪。",
        "targetPerson":  "「広がりを抑え、しっとりまとめたい」方。",
        "solvesConcern": "硬毛向けの保湿成分が髪をやわらかく整え、広がりを抑えてしっとりとしたまとまりを与えます。",
        "keyIngredients": [],
        "effects": [
            "硬毛・多毛をやわらかく整える",
            "広がり・パサつきを抑える",
            "しっとりとした質感を与える",
        ],
        "usage": "タオルドライ後、毛先〜中間に。",
    },
    ("エルジューダ", "エマルジョン"): {
        "pitchCopy":     "普通〜硬毛の毛先に、上品なやわらかさを。",
        "targetHair":    "普通毛〜硬毛で、毛先のまとまりを整えたい髪。",
        "targetPerson":  "「毛先のまとまりとやわらかさを叶えたい」方。",
        "solvesConcern": "ミルク質感がしっとりとなじみ、毛先のまとまりと指通りをなめらかに整えます。",
        "keyIngredients": [],
        "effects": [
            "毛先のまとまりとやわらかさを整える",
            "なめらかな指通りを与える",
            "毎日のアウトバスとして使いやすい",
        ],
        "usage": "タオルドライ後、毛先〜中間に。",
    },
    ("エルジューダ", "フリッズフィクサー エマルジョン"): {
        "pitchCopy":     "湿気とうねりに、フリッズケアという解決策。",
        "targetHair":    "湿気・うねり・広がりで膨らみやすい髪。",
        "targetPerson":  "「梅雨や湿気の日でもまとまる髪でいたい」方。",
        "solvesConcern": "湿気による広がりにアプローチするフリッズケア処方が、髪を扱いやすい質感へ整えます。",
        "keyIngredients": [],
        "effects": [
            "湿気・うねりによる広がりを抑える",
            "毛先のまとまりを整える",
            "扱いやすい質感へ導く",
        ],
        "usage": "タオルドライ後、毛先〜中間に。",
    },
    ("エルジューダ", "ジェルセラム"): {
        "pitchCopy":     "ブリーチ毛の絡まりに、ジェル発想の指通りケア。",
        "targetHair":    "ブリーチ・ハイトーンで絡まりやすい、ダメージ毛。",
        "targetPerson":  "「ハイトーンを楽しみたいが、絡まりやパサつきが気になる」方。",
        "solvesConcern": "ブリーチケアに特化したジェル&セラム処方が、絡まりを抑えてなめらかな指通りを与えます。",
        "keyIngredients": [],
        "effects": [
            "毛先の絡まりを抑える",
            "なめらかな指通りを与える",
            "ブリーチ毛のパサつきをケア",
        ],
        "usage": "タオルドライ後、毛先〜中間に。",
    },
    # ---------- Oggi otto ----------
    ("オッジィオット", "スキャルプ"): {
        "pitchCopy":     "頭皮を整え、髪の土台から健やかに。",
        "targetHair":    "頭皮環境を整えたい、根元ケアを始めたい髪。",
        "targetPerson":  "「頭皮の不調・根元の弱りが気になる」方。",
        "solvesConcern": "頭皮環境に着目した処方が地肌を健やかに整え、髪の土台からケアします。",
        "keyIngredients": [],
        "effects": [
            "頭皮環境を整える",
            "健やかな髪の土台を育む",
            "根元から扱いやすい髪へ",
        ],
    },
    ("オッジィオット", "セラム"): {
        "pitchCopy":     "オーダーメイド発想で、髪に必要な補修と保湿を。",
        "targetHair":    "髪質に合わせて細かく選びたいダメージ毛。",
        "targetPerson":  "「自分の髪に合うケアを精密に選びたい」方。",
        "solvesConcern": "CMC補修成分を中心とした処方が、髪質に応じた補修と保湿を実現します。",
        "keyIngredients": ["CMC"],
        "effects": [
            "髪の内部補修と保湿を両立",
            "髪質に合わせた質感調整",
            "なめらかな指通りを与える",
        ],
    },
    ("オッジィオット", "アウトバス"): {
        "pitchCopy":     "乾かす前後のケアで、仕上がりを底上げする。",
        "targetHair":    "アウトバスでも質感を妥協したくない髪。",
        "targetPerson":  "「乾かす前にもしっかりケアして、ツヤと指通りを整えたい」方。",
        "solvesConcern": "CMC配合のアウトバスが、ドライ前後の髪を補修・保湿し、毎日のスタイリングを引き上げます。",
        "keyIngredients": ["CMC"],
        "effects": [
            "乾かす前後の補修と保湿",
            "毛先のツヤと指通りを整える",
            "毎日の仕上げを底上げ",
        ],
        "usage": "タオルドライ後、毛先〜中間に。",
    },
    # ---------- Aujua ----------
    ("オージュア", "アクアヴィア"): {
        "pitchCopy":     "くせ・うねりに、毎日のまとまりを。",
        "targetHair":    "くせ・うねりによる広がり、湿気で膨らむ髪。",
        "targetPerson":  "「くせ毛で広がりやすい」「湿気の日もまとまっていたい」方。",
        "solvesConcern": "くせ・うねりにアプローチするオージュア独自処方が、扱いやすいまとまりへと整えます。",
        "keyIngredients": [],
        "effects": [
            "くせ・うねりによる広がりを抑える",
            "湿気による膨らみを整える",
            "扱いやすいまとまりを与える",
        ],
    },
    ("オージュア", "イミュライズ"): {
        "pitchCopy":     "エイジングダメージに、ハリのある補修を。",
        "targetHair":    "繰り返したカラー・パーマ、年齢によるダメージが気になる髪。",
        "targetPerson":  "「ダメージとエイジングの両方をケアしたい」方。",
        "solvesConcern": "毛髪強度に着目した処方が、エイジングによる弱りを整え、ハリのある補修感を与えます。",
        "keyIngredients": [],
        "effects": [
            "エイジングダメージにアプローチ",
            "毛髪強度を整え、ハリのある質感へ",
            "繰り返したダメージを補修",
        ],
    },
    ("オージュア", "インメトリィ"): {
        "pitchCopy":     "髪のゆがみを整え、ツヤとまとまりを引き出す。",
        "targetHair":    "うねり・広がり・エイジングによるまとまりにくさが気になる髪。",
        "targetPerson":  "「年齢によるうねりや、まとまりにくさが気になり始めた」方。",
        "solvesConcern": "髪のゆがみケアに特化した処方が、うねりを整え、ツヤとまとまりを引き出します。",
        "keyIngredients": [],
        "effects": [
            "髪のゆがみケアでうねりを整える",
            "ツヤとまとまりを引き出す",
            "扱いやすい質感へ導く",
        ],
    },
    ("オージュア", "エイジングスパ"): {
        "pitchCopy":     "頭皮から、ふんわり立ち上がる髪を。",
        "targetHair":    "トップのボリューム不足が気になり、頭皮ケアを始めたい髪。",
        "targetPerson":  "「ふんわり立ち上がる髪に戻りたい」方。",
        "solvesConcern": "地肌ケアに着目した処方が、頭皮を整え、根元からふんわり立ち上がる質感を引き出します。",
        "keyIngredients": [],
        "effects": [
            "頭皮環境を整える",
            "根元からふんわり立ち上がる質感",
            "エイジングケアの第一歩に",
        ],
    },
    ("オージュア", "クエンチ"): {
        "pitchCopy":     "髪の渇きに、しっとりとした答えを。",
        "targetHair":    "乾燥・パサつき・ごわつきが気になる髪。水分量が不足しがちな状態。",
        "targetPerson":  "「いろいろ試しても潤いが続かない」「毛先がカサつく」と感じている方。",
        "solvesConcern": "水分不足によるパサつき・ごわつき・まとまりにくさを、髪内部の水分保持力から立て直します。",
        "keyIngredients": [],
        "effects": [
            "髪の水分保持を高め、うるおいを持続",
            "やわらかく、しっとりまとまる質感へ",
            "繰り返しの乾燥にも継続的な保湿",
        ],
    },
    ("オージュア", "スムース"): {
        "pitchCopy":     "細毛・絡まりに、軽い指通りという答え。",
        "targetHair":    "細毛・軟毛・絡まり・指通りの悪さが気になる髪。",
        "targetPerson":  "「重くなりたくない」「さらさらの指通りが好き」な方。",
        "solvesConcern": "軽やかな仕上がりを叶える処方が、絡まりを抑え、さらさらの指通りを引き出します。",
        "keyIngredients": [],
        "effects": [
            "絡まりを抑え、なめらかな指通りに",
            "細毛・軟毛をふんわり仕上げる",
            "軽さと扱いやすさを両立",
        ],
    },
    ("オージュア", "フィルメロウ"): {
        "pitchCopy":     "アイロン・コテで硬くなった髪に、やわらかさを。",
        "targetHair":    "毎日のアイロン・コテ・ブローによる硬さやごわつきが気になる髪。",
        "targetPerson":  "「アイロンの熱ダメージで髪が硬くなりがち」な方。",
        "solvesConcern": "熱ダメージケアに着目した処方が、硬くなった髪をやわらかく整え、まとまりを引き出します。",
        "keyIngredients": [],
        "effects": [
            "熱ダメージによる硬さをやわらげる",
            "やわらかさとまとまりを両立",
            "毎日の熱スタイリングを支える",
        ],
    },
    ("オージュア", "モイストカーム"): {
        "pitchCopy":     "頭皮の乾燥・ゆらぎに、いたわりの保湿を。",
        "targetHair":    "頭皮の乾燥・かゆみ・季節のゆらぎが気になる髪。",
        "targetPerson":  "「乾燥する季節は頭皮もゆらぎがち」と感じる方。",
        "solvesConcern": "頭皮の保湿に着目した処方が、敏感に傾きがちな地肌をいたわり、健やかに整えます。",
        "keyIngredients": [],
        "effects": [
            "頭皮を保湿し、いたわる",
            "乾燥・かゆみ・ゆらぎをケア",
            "敏感頭皮にも配慮した使用感",
        ],
    },
    ("オージュア", "リペアリティ"): {
        "pitchCopy":     "ブリーチ毛の弱りに、しなやかさを取り戻す。",
        "targetHair":    "ブリーチ毛・ハイトーン・切れ毛やごわつきが気になる髪。",
        "targetPerson":  "「ブリーチを楽しみたいが、髪のダメージはしっかりケアしたい」方。",
        "solvesConcern": "ハイダメージ毛の補修に特化した処方が、内部の弱りを整え、しなやかな質感を取り戻します。",
        "keyIngredients": [],
        "effects": [
            "ハイダメージ毛を内部から補修",
            "切れ毛・ごわつきを抑える",
            "しなやかでやわらかい質感へ",
        ],
    },
    # ---------- Global Milbon ----------
    ("グローバルミルボン", "アンチフリッズ"): {
        "pitchCopy":     "湿気・うねりに、グローバル発想のフリッズケアを。",
        "targetHair":    "くせ・うねり・湿気でまとまらない髪。",
        "targetPerson":  "「湿気でいつも髪が広がる」方。",
        "solvesConcern": "湿気と広がりに着目した処方が、髪をなめらかに整え、扱いやすいまとまりを与えます。",
        "keyIngredients": [],
        "effects": [
            "湿気・うねりによる広がりを抑える",
            "なめらかなまとまりを与える",
            "扱いやすい質感へ整える",
        ],
    },
    ("グローバルミルボン", "ブロンドプラス"): {
        "pitchCopy":     "ブリーチ毛の黄ばみと透明感に。",
        "targetHair":    "ブリーチ毛・ハイトーンカラーの黄ばみが気になる髪。",
        "targetPerson":  "「ハイトーンの透明感をきれいに保ちたい」方。",
        "solvesConcern": "黄ばみケアに特化した処方が、ハイトーンカラーの透明感とツヤを保ちます。",
        "keyIngredients": [],
        "effects": [
            "黄ばみを抑え、透明感を保つ",
            "ブリーチ毛の質感を整える",
            "ハイトーンカラーをきれいに",
        ],
    },
    ("グローバルミルボン", "カラープリザーブ"): {
        "pitchCopy":     "カラーの色持ちに、ツヤと安心を。",
        "targetHair":    "カラーの褪色やカラー後の乾燥が気になる髪。",
        "targetPerson":  "「カラーを長くきれいに楽しみたい」方。",
        "solvesConcern": "カラーケア設計が褪色を抑え、ツヤと色持ちを両立します。",
        "keyIngredients": [],
        "effects": [
            "カラーの褪色を抑える",
            "色持ちとツヤを両立",
            "カラー後の乾燥もケア",
        ],
    },
    ("グローバルミルボン", "モイスチュア"): {
        "pitchCopy":     "乾燥・パサつきに、しっとりとした答えを。",
        "targetHair":    "乾燥・パサつき・毛先のまとまり不足が気になる髪。",
        "targetPerson":  "「うるおいで満たして、しっとりまとめたい」方。",
        "solvesConcern": "保湿に着目した処方が、乾燥した髪に水分を満たし、しっとりとしたまとまりを与えます。",
        "keyIngredients": [],
        "effects": [
            "髪に水分を満たし、うるおいを与える",
            "毛先までしっとりまとめる",
            "乾燥による広がりを抑える",
        ],
    },
    ("グローバルミルボン", "リペア"): {
        "pitchCopy":     "ダメージに、しなやかな補修を。",
        "targetHair":    "カラー・パーマなどによるダメージが気になる髪。",
        "targetPerson":  "「ダメージケアをきちんと続けたい」方。",
        "solvesConcern": "ダメージ補修処方が髪内部にアプローチし、しなやかさと指通りを取り戻します。",
        "keyIngredients": [],
        "effects": [
            "ダメージを内部から補修",
            "しなやかな質感と指通りを与える",
            "扱いやすいまとまりへ整える",
        ],
    },
    ("グローバルミルボン", "リペアヒート"): {
        "pitchCopy":     "アイロン熱に、強い髪を育てる。",
        "targetHair":    "毎日のアイロン・コテで硬くなりがちな髪。",
        "targetPerson":  "「熱スタイリングを毎日続けるが、髪は守りたい」方。",
        "solvesConcern": "熱ダメージにアプローチする処方が、アイロンの熱から髪を守り、なめらかさを引き出します。",
        "keyIngredients": [],
        "effects": [
            "アイロン熱から髪を守る",
            "熱ダメージをケアし、なめらかな質感へ",
            "扱いやすい指通りを与える",
        ],
    },
    ("グローバルミルボン", "スムース"): {
        "pitchCopy":     "絡まりに、軽やかなさらさら感を。",
        "targetHair":    "細毛・絡まりが気になり、軽い仕上がりを好む髪。",
        "targetPerson":  "「重くしたくない」「さらさらの仕上がりが好き」な方。",
        "solvesConcern": "軽やかな仕上がりを叶える処方が、絡まりを抑え、さらさらの指通りを引き出します。",
        "keyIngredients": [],
        "effects": [
            "絡まりを抑え、なめらかな指通りに",
            "軽い質感とまとまりを両立",
            "細毛・軟毛をふんわり仕上げる",
        ],
    },
    # ---------- Kérastase ----------
    ("ケラスターゼ", "ブロンドアブソリュ"): {
        "pitchCopy":     "ハイトーンの透明感を、毎日に。",
        "targetHair":    "ブリーチ毛・ハイトーンカラーをきれいに保ちたい髪。",
        "targetPerson":  "「ブリーチの透明感をできるだけ長く楽しみたい」方。",
        "solvesConcern": "ブロンド毛の黄ばみケアに特化した処方が、透明感と上質なツヤを保ちます。",
        "keyIngredients": [],
        "effects": [
            "ブロンド毛の黄ばみを抑える",
            "透明感とツヤを保つ",
            "ハイトーン毛の質感を整える",
        ],
    },
    ("ケラスターゼ", "クロマアブソリュ"): {
        "pitchCopy":     "カラーの輝きを、最大限引き出す。",
        "targetHair":    "カラーの色持ちやツヤを重視する髪。",
        "targetPerson":  "「カラーの色持ちにこだわりたい」方。",
        "solvesConcern": "カラーケア処方が褪色を防ぎ、ツヤやかな色持ちを支えます。",
        "keyIngredients": [],
        "effects": [
            "カラーの褪色を抑える",
            "色持ちとツヤを最大限に",
            "カラー毛の質感を整える",
        ],
    },
    ("ケラスターゼ", "クロノロジスト"): {
        "pitchCopy":     "ケラスターゼ最高峰。総合的に髪を底上げする。",
        "targetHair":    "髪全体の印象を底上げしたい、ギフトにもふさわしい上質ライン。",
        "targetPerson":  "「最上級の体験を、自分や大切な人に贈りたい」方。",
        "solvesConcern": "ブランド最高峰の処方が、ツヤ・補修・保湿を総合的に底上げします。",
        "keyIngredients": [],
        "effects": [
            "髪全体の印象を上質に底上げ",
            "ツヤと指通りを引き出す",
            "ギフトにも選ばれる上質感",
        ],
    },
    ("ケラスターゼ", "ディシプリン"): {
        "pitchCopy":     "うねり・広がりに、規律あるまとまりを。",
        "targetHair":    "うねり・広がり・湿気でまとまらない髪。",
        "targetPerson":  "「くせ毛・うねりをすっきりまとめたい」方。",
        "solvesConcern": "うねりにアプローチする処方が、髪を規律あるなめらかさへと整えます。",
        "keyIngredients": [],
        "effects": [
            "うねり・広がりを抑える",
            "なめらかなまとまりを与える",
            "湿気の影響を受けにくい質感へ",
        ],
    },
    ("ケラスターゼ", "ジェネシス"): {
        "pitchCopy":     "切れ毛・髪の弱りに、しなやかな強さを。",
        "targetHair":    "抜け毛・切れ毛・髪の弱りが気になる髪。",
        "targetPerson":  "「髪が細く、弱ってきた」と感じる方。",
        "solvesConcern": "切れ毛ケアに着目した処方が、髪を内側からしなやかに整え、弱りにアプローチします。",
        "keyIngredients": [],
        "effects": [
            "切れ毛・抜け毛にアプローチ",
            "しなやかで強い質感へ整える",
            "髪の弱りをケア",
        ],
    },
    ("ケラスターゼ", "ニュートリティブ"): {
        "pitchCopy":     "乾燥した髪に、深いうるおいを。",
        "targetHair":    "乾燥・パサつき・ゴワつきが気になる髪。",
        "targetPerson":  "「しっとりやわらかい質感を毎日続けたい」方。",
        "solvesConcern": "保湿に特化した処方が、乾燥した髪をやわらかく整え、ツヤを引き出します。",
        "keyIngredients": [],
        "effects": [
            "乾燥した髪に深いうるおいを与える",
            "やわらかさとツヤを両立",
            "ゴワつきを抑え、扱いやすい質感へ",
        ],
    },
    ("ケラスターゼ", "レジスタンス"): {
        "pitchCopy":     "ダメージに、補修と強さを。",
        "targetHair":    "ダメージ・切れ毛・枝毛が気になる髪。",
        "targetPerson":  "「カラーやパーマで弱った髪を立て直したい」方。",
        "solvesConcern": "ダメージレベルに応じた補修処方が、髪の強さとしなやかさを取り戻します。",
        "keyIngredients": [],
        "effects": [
            "ダメージを内部から補修",
            "切れ毛・枝毛にアプローチ",
            "強さとしなやかさを両立",
        ],
    },
    # ---------- SUBLIMIC ----------
    ("サブリミック", "アクアインテンシブ"): {
        "pitchCopy":     "乾燥・カラー後の弱りに、しなやかな保湿を。",
        "targetHair":    "乾燥・パサつき・カラーやパーマ後の弱りが気になる髪。",
        "targetPerson":  "「カラー後の乾燥としなやかさの低下を立て直したい」方。",
        "solvesConcern": "サブリミック発想の保湿処方が、乾燥した髪をしなやかに整え、うるおいを与えます。",
        "keyIngredients": [],
        "effects": [
            "乾燥した髪にうるおいを与える",
            "しなやかさを取り戻す",
            "カラー後の弱りをケア",
        ],
    },
    ("サブリミック", "アデノバイタル"): {
        "pitchCopy":     "ハリコシ・ボリュームに、年齢を超える土台ケア。",
        "targetHair":    "細毛・ハリコシ不足・年齢によるボリューム悩みが気になる髪。",
        "targetPerson":  "「ハリコシとボリュームを取り戻したい」方。",
        "solvesConcern": "ハリコシケアに特化した処方が、髪と頭皮の両方を整え、立ち上がるボリュームを引き出します。",
        "keyIngredients": [],
        "effects": [
            "ハリコシのある質感へ整える",
            "頭皮ケアでボリューム不足にアプローチ",
            "エイジング世代の髪の土台を支える",
        ],
    },
    ("サブリミック", "エアリーフロー"): {
        "pitchCopy":     "くせ・広がりに、軽やかなまとまりを。",
        "targetHair":    "くせ・広がりが気になり、重くしたくない髪。",
        "targetPerson":  "「広がりは抑えたいけれど、軽い仕上がりがいい」方。",
        "solvesConcern": "軽やかさとまとまりを両立する処方が、くせ・広がりを抑えながら扱いやすい質感へ整えます。",
        "keyIngredients": [],
        "effects": [
            "くせ・広がりを抑える",
            "軽やかな仕上がりを与える",
            "重くなりがちな髪を扱いやすく",
        ],
    },
    ("サブリミック", "フェンテフォルテ"): {
        "pitchCopy":     "頭皮の不調に、地肌からのリセットを。",
        "targetHair":    "頭皮のべたつき・乾燥・不安定さが気になる髪。",
        "targetPerson":  "「頭皮環境を立て直したい」方。",
        "solvesConcern": "頭皮ケアに特化した処方が、地肌環境を整え、すっきりとした洗い心地を与えます。",
        "keyIngredients": [],
        "effects": [
            "頭皮環境を整える",
            "べたつき・乾燥にアプローチ",
            "すっきりとした洗い心地を与える",
        ],
    },
    ("サブリミック", "ルミノフォース"): {
        "pitchCopy":     "カラーのツヤと色持ちに、磨きをかける。",
        "targetHair":    "カラーの褪色や色ツヤが気になる髪。",
        "targetPerson":  "「カラーをきれいに長く楽しみたい」方。",
        "solvesConcern": "カラーケア設計が褪色を防ぎ、ツヤやかな色持ちを支えます。",
        "keyIngredients": [],
        "effects": [
            "カラーの褪色を抑える",
            "ツヤと色持ちを両立",
            "カラー毛の質感を整える",
        ],
    },
    ("サブリミック", "ワンダーシールド"): {
        "pitchCopy":     "日常のダメージから、髪を守る一本。",
        "targetHair":    "湿気・紫外線・摩擦など日常ダメージが気になる髪。",
        "targetPerson":  "「サロン仕上がりを毎日続けたい」方。",
        "solvesConcern": "外的ストレスにアプローチする処方が、サロン帰りのなめらかさを毎日持続させます。",
        "keyIngredients": [],
        "effects": [
            "湿気・摩擦・紫外線などの外的ストレスをケア",
            "サロン仕上がりのなめらかさを持続",
            "毎日扱いやすい質感へ整える",
        ],
        "usage": "タオルドライ後、毛先〜中間に。",
    },
    # ---------- SPRINAGE ----------
    ("スプリナージュ", "SPRINAGE HAIR CARE"): {
        "pitchCopy":     "やさしい使用感で、髪と頭皮を整える。",
        "targetHair":    "やさしい使用感を求める髪。",
        "targetPerson":  "「自然派の印象で、やさしいケアを続けたい」方。",
        "solvesConcern": "やさしい使用感の処方が、髪と頭皮をいたわりながら、保湿と地肌ケアを支えます。",
        "keyIngredients": [],
        "effects": [
            "髪と頭皮をやさしくいたわる",
            "うるおいと地肌ケアを両立",
            "自然派の印象に寄り添う使用感",
        ],
    },
    # ---------- track ----------
    ("トラック", "track Cream"): {
        "pitchCopy":     "自然なまとまりと束感を、クリームで自在に。",
        "targetHair":    "自然な束感とまとまりを求める仕上げ前の髪。",
        "targetPerson":  "「自然な束感とまとまりを毎日楽しみたい」方。",
        "solvesConcern": "クリーム発想の質感調整が、髪に自然なまとまりと束感を与えます。",
        "keyIngredients": [],
        "effects": [
            "自然なまとまりと束感を演出",
            "仕上げの質感を整える",
            "扱いやすいスタイリング性",
        ],
        "usage": "乾いた髪の仕上げ、または乾かす前のセット前に。",
    },
    # ---------- NANOAMINO ----------
    ("ナノアミノ", "NANOAMINO HOMECARE"): {
        "pitchCopy":     "毎日のホームケアに、補修と保湿を。",
        "targetHair":    "日々のケアで補修と保湿を整えたい髪。",
        "targetPerson":  "「毎日のシャンプー&トリートメントで手触りを整えたい」方。",
        "solvesConcern": "ホームケア発想の処方が、日々のケアに補修と保湿を加え、扱いやすい質感へ整えます。",
        "keyIngredients": [],
        "effects": [
            "日々のケアに補修と保湿を加える",
            "手触りを整え、まとまりを与える",
            "毎日続けやすい使用感",
        ],
    },
    # ---------- HAHONICO ----------
    ("ハホニコ", "キラメラメ系"): {
        "pitchCopy":     "サロンの集中ケア発想で、毛先までツヤと手触りを。",
        "targetHair":    "集中補修やサロンケア感を求める髪。",
        "targetPerson":  "「サロン帰りの手触りとツヤを家でも続けたい」方。",
        "solvesConcern": "集中補修発想の処方が、ツヤと手触りを整え、サロン仕上がりに近づけます。",
        "keyIngredients": [],
        "effects": [
            "毛先まで集中的に補修",
            "ツヤと手触りを引き出す",
            "サロンケア発想の仕上がり",
        ],
        "usage": "週1〜2回の集中ケアとして。",
    },
    # ---------- Vie qualite ----------
    ("バイカルテ", "リペアシャンプー CH+"): {
        "pitchCopy":     "普通毛〜硬毛に、まとまりのある補修を。",
        "targetHair":    "普通毛〜硬毛でダメージ補修とまとまりを求める髪。",
        "targetPerson":  "「ダメージを補修しながら、しっかりまとめたい」方。",
        "solvesConcern": "普通毛〜硬毛向けの補修処方が、髪を内側から整え、まとまりとなめらかさを与えます。",
        "keyIngredients": [],
        "effects": [
            "普通毛〜硬毛をしっかり補修",
            "まとまりとなめらかさを与える",
            "扱いやすい質感へ整える",
        ],
    },
    ("バイカルテ", "リペアシャンプー FH+"): {
        "pitchCopy":     "細毛・軟毛に、軽さを残した補修を。",
        "targetHair":    "細毛・軟毛で補修したいが重くしたくない髪。",
        "targetPerson":  "「補修したいけれど、軽さは残したい」方。",
        "solvesConcern": "細毛・軟毛向けの補修処方が、軽さを保ちながら髪を整え、ふんわりとしたまとまりを与えます。",
        "keyIngredients": [],
        "effects": [
            "細毛・軟毛を軽やかに補修",
            "ふんわりとしたまとまりを与える",
            "重さを残さず質感を整える",
        ],
    },
    ("バイカルテ", "リペアシャンプー UH+"): {
        "pitchCopy":     "くせ・うねりに、補修とまとまりを。",
        "targetHair":    "くせ毛・うねり・広がりが気になる髪。",
        "targetPerson":  "「うねりを抑えながら補修もしたい」方。",
        "solvesConcern": "うねりケアに着目した補修処方が、髪をなめらかにまとめ、扱いやすい質感へ整えます。",
        "keyIngredients": [],
        "effects": [
            "うねり・広がりを抑える",
            "ダメージ補修とまとまりを両立",
            "扱いやすい質感へ整える",
        ],
    },
    ("バイカルテ", "セラムトリートメント SS+/MS+/HS+"): {
        "pitchCopy":     "髪質に合わせて、補修と質感を選ぶ。",
        "targetHair":    "髪質に合わせて補修と仕上がりを選びたい髪。",
        "targetPerson":  "「自分の髪質に合うトリートメントを精密に選びたい」方。",
        "solvesConcern": "髪質別に設計された補修処方が、ダメージを整え、希望の仕上がりへ導きます。",
        "keyIngredients": [],
        "effects": [
            "髪質に合わせた補修",
            "なめらかな指通りを与える",
            "仕上がりの質感を選べる",
        ],
    },
    ("バイカルテ", "コンセントレイトエッセンス"): {
        "pitchCopy":     "毛先の弱りに、集中ケアの一手。",
        "targetHair":    "毛先のダメージや弱りが気になる髪。",
        "targetPerson":  "「毛先を集中的に補修したい」方。",
        "solvesConcern": "毛先専用の補修美容液が、なめらかさとまとまりを集中的に与えます。",
        "keyIngredients": [],
        "effects": [
            "毛先を集中的に補修",
            "なめらかな質感を取り戻す",
            "まとまりやすい毛先に整える",
        ],
        "usage": "週1〜2回、毛先を中心に。",
    },
    ("バイカルテ", "セラムマスク"): {
        "pitchCopy":     "毛羽立つ毛先に、週1-2回の集中補修を。",
        "targetHair":    "毛羽立ち・ダメージが気になる毛先。",
        "targetPerson":  "「毛先のダメージを集中的に整えたい」方。",
        "solvesConcern": "集中補修処方が、毛先のダメージを整え、なめらかな質感を取り戻します。",
        "keyIngredients": [],
        "effects": [
            "毛先のダメージを集中的に補修",
            "毛羽立ちを抑え、なめらかに整える",
            "週1〜2回のスペシャルケアに",
        ],
        "usage": "週1〜2回、毛先を中心に。",
    },
    ("バイカルテ", "エッセンスミルク"): {
        "pitchCopy":     "毛先のパサつきに、しっとりまとめるアウトバスを。",
        "targetHair":    "毛先のパサつきや広がりが気になる髪。",
        "targetPerson":  "「毛先のまとまりとやわらかさを整えたい」方。",
        "solvesConcern": "アウトバス補修処方が、毛先のパサつきと広がりを整え、しっとりまとめます。",
        "keyIngredients": [],
        "effects": [
            "毛先のパサつき・広がりを抑える",
            "しっとりとしたまとまりを与える",
            "やわらかな指通りに整える",
        ],
        "usage": "タオルドライ後、毛先〜中間に。",
    },
    # ---------- LIKERA ----------
    ("リケラ", "リケラエマルジョン"): {
        "pitchCopy":     "熱ダメージに、しなやかな補修を。",
        "targetHair":    "アイロン・コテ・ドライヤーをよく使う、熱ダメージが気になる髪。",
        "targetPerson":  "「毎日のヒートスタイリングをしながら、髪のしなやかさは守りたい」方。",
        "solvesConcern": "熱ダメージにアプローチする処方が、髪をしっとりしなやかに整え、扱いやすさを取り戻します。",
        "keyIngredients": [],
        "effects": [
            "熱ダメージを補修",
            "しっとりしなやかな質感へ整える",
            "毎日のスタイリングを支える",
        ],
    },
    # ---------- 髪にドラマを。 ----------
    ("髪にドラマを。", "つるりんちょ。シャンプー"): {
        "pitchCopy":     "縮毛矯正・うねりの広がりに、つるんと素直な髪を。",
        "targetHair":    "縮毛矯正毛・くせ毛・広がりが気になる髪。",
        "targetPerson":  "「縮毛矯正後の質感やうねりの広がりを整えたい」方。",
        "solvesConcern": "うねりケアに着目した処方が、縮毛矯正毛にもなじみ、つるんとまとまる質感を引き出します。",
        "keyIngredients": [],
        "effects": [
            "うねり・広がりを抑える",
            "縮毛矯正毛にもなじむまとまり",
            "つるんとした指通りを与える",
        ],
    },
    ("髪にドラマを。", "つるりんちょ。ブースター"): {
        "pitchCopy":     "質感を底上げする、集中ケアのブースター。",
        "targetHair":    "質感を底上げしたい、なめらかさが欲しい髪。",
        "targetPerson":  "「日々のケアの仕上がりをワンランク引き上げたい」方。",
        "solvesConcern": "集中補修発想のブースターが、髪の質感を引き上げ、まとまりとなめらかさを与えます。",
        "keyIngredients": [],
        "effects": [
            "髪の質感を底上げ",
            "なめらかさとまとまりを与える",
            "日々のケアの仕上がりを引き上げる",
        ],
    },
    ("髪にドラマを。", "つるりんちょ。トリートメント"): {
        "pitchCopy":     "毛先まで、つるんとまとまる仕上がりへ。",
        "targetHair":    "乾燥や広がりが気になる髪。",
        "targetPerson":  "「毛先まで均一にまとまる髪でいたい」方。",
        "solvesConcern": "なめらかな質感を引き出す処方が、毛先まで均一にまとめ、扱いやすい質感へ整えます。",
        "keyIngredients": [],
        "effects": [
            "毛先までなめらかにまとめる",
            "乾燥・広がりを抑える",
            "指通りを整える",
        ],
    },
    ("髪にドラマを。", "つるりんちょ。トリートメント SARARITO"): {
        "pitchCopy":     "細毛・軽さ重視に、さらりとした仕上がりを。",
        "targetHair":    "細毛・軽い仕上がりを好む髪。",
        "targetPerson":  "「重くなりたくない」「さらっと軽い仕上がりが好き」な方。",
        "solvesConcern": "軽やかな質感を叶える処方が、絡まりを抑え、さらりとした指通りに整えます。",
        "keyIngredients": [],
        "effects": [
            "軽やかでさらりとした仕上がり",
            "細毛・軟毛をふんわり整える",
            "絡まりを抑え、指通りを良くする",
        ],
    },
    ("髪にドラマを。", "いるかのせなか。ベースケア"): {
        "pitchCopy":     "ドライ前のベースケアで、仕上がりを整える。",
        "targetHair":    "日々のドライ前ケアを整えたい髪。",
        "targetPerson":  "「乾かす前のケアを整え、仕上がりを引き上げたい」方。",
        "solvesConcern": "ベースケア発想のアウトバスが、ドライ前の髪を整え、指通りとまとまりを与えます。",
        "keyIngredients": [],
        "effects": [
            "ドライ前の髪を整える",
            "指通りとまとまりを与える",
            "毎日のスタイリングを引き上げる",
        ],
        "usage": "タオルドライ後、毛先〜中間に。",
    },
}


# ============================================================
# 7) カテゴリ別の usage デフォルト
# ============================================================
def default_usage(category):
    return {
        "shampoo":        "毎日のシャンプーとして。",
        "treatment":      "シャンプー後、毛先〜中間になじませて流す。",
        "hair-mask":      "週1〜2回、シャンプー後の集中ケアとして。",
        "out-bath-milk":  "タオルドライ後、毛先〜中間に。",
        "out-bath-oil":   "タオルドライ後、または仕上げに毛先中心に。",
        "scalp-essence":  "夜、タオルドライ後の頭皮に。マッサージしながら。",
        "styling":        "乾いた髪の仕上げに、毛先中心に少量。",
        "heat-protect":   "アイロン・コテの前に。タオルドライ後の髪に。",
        "beauty-tool":    "取扱説明書に従って使用。",
    }.get(category, "")


# ============================================================
# 8) ID 生成
# ============================================================
def make_id(brand_slug, line_slug, role, brand_jp, line_jp, name):
    base = f"{brand_slug}-{line_slug}-{role}"
    # 同ブランド・同ライン・同カテゴリで複数商品があるケースに名前から接尾辞を付ける
    return base, name


# ============================================================
# 9) メイン処理
# ============================================================
def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["公式特徴追加_商品一覧"]
    headers = [c.value for c in ws[1]]

    rows = []
    for ri in range(2, ws.max_row + 1):
        row = [c.value for c in ws[ri]]
        rd = dict(zip(headers, row))
        rows.append(rd)

    # 同 (brand, line, role) で複数商品があるかを先に確認
    seen_id_count = {}

    products = []
    for rd in rows:
        brand_jp = rd.get("brand_jp")
        line_jp  = rd.get("line_jp")
        cat_jp   = rd.get("ai_category_simple")
        name     = rd.get("product_name")
        points   = rd.get("official_feature_points")
        recom    = rd.get("official_recommended_for")
        summary  = rd.get("official_feature_summary")
        role     = rd.get("category_role_feature")
        image_fn = rd.get("download_filename_suggestion")

        if not (brand_jp and cat_jp and name):
            continue

        bmap = BRAND_MAP.get(brand_jp)
        if not bmap:
            continue
        line_slug = LINE_SLUG.get((brand_jp, line_jp), "core")

        category, routine, cat_short = detect_category(cat_jp, name)
        base_id, _ = make_id(bmap["slug"], line_slug, cat_short, brand_jp, line_jp, name)

        # 重複ID対応：同基底IDが既にある場合は連番を付ける
        seen_id_count[base_id] = seen_id_count.get(base_id, 0) + 1
        pid = base_id if seen_id_count[base_id] == 1 else f"{base_id}-{seen_id_count[base_id]}"

        tags = derive_tags(brand_jp, line_jp, name, cat_jp, points, recom, category, role)

        # ストーリー層
        narr = LINE_NARRATIVE.get((brand_jp, line_jp), {})
        pitchCopy     = narr.get("pitchCopy")
        targetHair    = narr.get("targetHair")
        targetPerson  = narr.get("targetPerson")
        solvesConcern = narr.get("solvesConcern")
        keyIngredients = narr.get("keyIngredients", [])
        effects       = narr.get("effects", [])
        usage         = narr.get("usage") or default_usage(category)

        obj = {
            "id":          pid,
            "brand":       bmap["display"],
            "maker":       bmap["maker"],
            "line":        line_jp or "",
            "name":        name,
            "category":    category,
            "routineStep": routine,
        }
        if image_fn:
            obj["image"] = f"images/brands/{image_fn}"

        # タグ（空配列は省略 — 推測しない原則）
        if tags["concernTags"]:
            obj["concernTags"] = tags["concernTags"]
        obj["hairType"]    = tags["hairType"]
        obj["damageTags"]  = tags["damageTags"]
        obj["finishTags"]  = tags["finishTags"]
        if tags["functionTags"]:
            obj["functionTags"] = tags["functionTags"]
        if tags["scalpFit"]:
            obj["scalpFit"] = tags["scalpFit"]
        obj["lengthFit"] = "all"
        if tags["preferenceFit"]:
            obj["preferenceFit"] = tags["preferenceFit"]
        if tags["usageTiming"]:
            obj["usageTiming"] = tags["usageTiming"]
        if tags["specialTech"]:
            obj["specialTech"] = tags["specialTech"]

        # ストーリー
        if pitchCopy:     obj["pitchCopy"]     = pitchCopy
        if targetHair:    obj["targetHair"]    = targetHair
        if targetPerson:  obj["targetPerson"]  = targetPerson
        if solvesConcern: obj["solvesConcern"] = solvesConcern
        if keyIngredients:
            obj["keyIngredients"] = keyIngredients
        if effects:
            obj["effects"] = effects
        if usage:
            obj["usage"] = usage

        # 参考情報（香り関連語を含むサマリーは除外 — 香り情報出力禁止）
        SCENT_WORDS = ["香り", "香気", "フレグランス", "調香", "アロマ"]
        if summary and not any(w in summary for w in SCENT_WORDS):
            obj["officialSummary"] = summary
        if rd.get("feature_source_url"):
            obj["sourceUrl"] = rd.get("feature_source_url")

        products.append(obj)

    Path(OUT_PATH).parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(products)} products to {OUT_PATH}")

    # サマリ表示
    from collections import Counter
    by_brand = Counter(p["brand"] for p in products)
    by_cat = Counter(p["category"] for p in products)
    print("\n--- by brand ---")
    for k, v in by_brand.most_common():
        print(f"  {k}: {v}")
    print("\n--- by category ---")
    for k, v in by_cat.most_common():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
